"""Import cameras from Cyanview Compatible Cameras Excel file."""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from clorag.core.database import get_camera_database
from clorag.models.camera import CameraCreate, CameraSource, DeviceType, normalize_camera_create

EXCEL_PATH = Path(
    "/Users/alanogic/ddev/cyanview-support/docs/integrations/"
    "Cyanview_Compatible_Cameras.xlsx"
)

# ─── Device type mapping: (manufacturer, model_substring) → DeviceType ───
# Order matters: first match wins within each manufacturer.
DEVICE_TYPE_MAP: dict[tuple[str, str], DeviceType] = {
    # AJA
    ("AJA", "RovoCam"): DeviceType.CAMERA_PTZ,
    # BirdDog — all PTZ
    ("BirdDog", ""): DeviceType.CAMERA_PTZ,
    # Bolin
    ("Bolin", "R9"): DeviceType.CAMERA_PTZ,
    # Blackmagic
    ("Blackmagic", "ATEM"): DeviceType.SWITCHER,
    ("Blackmagic", "Focus Handle"): DeviceType.OTHER,
    ("Blackmagic", "Zoom Handle"): DeviceType.OTHER,
    ("Blackmagic", "Pocket Cinema"): DeviceType.CAMERA_CINEMA,
    ("Blackmagic", "Cinema Camera 6K"): DeviceType.CAMERA_CINEMA,
    ("Blackmagic", "PYXIS 6K"): DeviceType.CAMERA_BOX,
    ("Blackmagic", "Micro Studio 4K G2"): DeviceType.CAMERA_BOX,
    ("Blackmagic", "Micro Studio Camera 4K G2"): DeviceType.CAMERA_BOX,
    ("Blackmagic", "Micro Studio 4K"): DeviceType.CAMERA_BOX,
    ("Blackmagic", "Micro Studio Camera 4K"): DeviceType.CAMERA_BOX,
    ("Blackmagic", "Studio Camera 6K Pro"): DeviceType.CAMERA_BROADCAST,
    ("Blackmagic", "Studio Camera 4K Plus G2"): DeviceType.CAMERA_BROADCAST,
    ("Blackmagic", "Studio Camera 4K Plus"): DeviceType.CAMERA_BROADCAST,
    ("Blackmagic", "Studio Camera 4K Pro G2"): DeviceType.CAMERA_BROADCAST,
    ("Blackmagic", "Studio Camera 4K Pro"): DeviceType.CAMERA_BROADCAST,
    ("Blackmagic", "URSA"): DeviceType.CAMERA_CINEMA,
    # Blackmagic protocol-only rows (IP, SDI)
    ("Blackmagic", "IP"): DeviceType.OTHER,
    ("Blackmagic", "SDI"): DeviceType.OTHER,
    # Canon
    ("Canon", "CR-N"): DeviceType.CAMERA_PTZ,
    ("Canon", "CR-X"): DeviceType.CAMERA_PTZ,
    ("Canon", "R5"): DeviceType.CAMERA_MIRRORLESS,
    ("Canon", "XF"): DeviceType.CAMERA_BROADCAST,
    ("Canon", "C"): DeviceType.CAMERA_CINEMA,  # C70, C100, C200, C300, C400, C500, C700, C80
    # Dreamchip — all module
    ("Dreamchip", ""): DeviceType.CAMERA_MODULE,
    # Ikegami — all broadcast
    ("Ikegami", ""): DeviceType.CAMERA_BROADCAST,
    # JVC
    ("JVC", ""): DeviceType.CAMERA_BROADCAST,
    # Marshall — all box
    ("Marshall", ""): DeviceType.CAMERA_BOX,
    # Panasonic
    ("Panasonic", "AU-EVA"): DeviceType.CAMERA_CINEMA,
    ("Panasonic", "VariCam"): DeviceType.CAMERA_CINEMA,
    ("Panasonic", "AW-HE"): DeviceType.CAMERA_PTZ,
    ("Panasonic", "AW-UE"): DeviceType.CAMERA_PTZ,
    ("Panasonic", "AW-UB"): DeviceType.CAMERA_BOX,
    ("Panasonic", "BS1H"): DeviceType.CAMERA_BOX,
    ("Panasonic", "DC-BGH1"): DeviceType.CAMERA_BOX,
    ("Panasonic", "DC-GH"): DeviceType.CAMERA_MIRRORLESS,
    ("Panasonic", "AG-CX"): DeviceType.CAMERA_BROADCAST,
    ("Panasonic", "AG-HPX"): DeviceType.CAMERA_BROADCAST,
    ("Panasonic", "AJ-HPX"): DeviceType.CAMERA_BROADCAST,
    ("Panasonic", "AJ-PX"): DeviceType.CAMERA_BROADCAST,
    ("Panasonic", "AK-"): DeviceType.CAMERA_BROADCAST,
    # Proton — all module
    ("Proton", ""): DeviceType.CAMERA_MODULE,
    # RED
    ("RED", "Komodo"): DeviceType.CAMERA_CINEMA,
    ("RED", "V-Raptor"): DeviceType.CAMERA_CINEMA,
    # Sony
    ("Sony", "Alpha"): DeviceType.CAMERA_MIRRORLESS,
    ("Sony", "ZV-E"): DeviceType.CAMERA_MIRRORLESS,
    ("Sony", "BRC"): DeviceType.CAMERA_PTZ,
    ("Sony", "FR7"): DeviceType.CAMERA_PTZ,
    ("Sony", "Burano"): DeviceType.CAMERA_CINEMA,
    ("Sony", "CBM"): DeviceType.CAMERA_CINEMA,
    ("Sony", "FCB"): DeviceType.CAMERA_MODULE,
    ("Sony", "FX"): DeviceType.CAMERA_CINEMA,
    ("Sony", "Venice"): DeviceType.CAMERA_CINEMA,
    ("Sony", "LANC"): DeviceType.CAMERA_CINEMA,
    ("Sony", "RX0"): DeviceType.CAMERA_ACTION,
    ("Sony", "HXR-NX"): DeviceType.CAMERA_BROADCAST,
    ("Sony", "PXW-Z"): DeviceType.CAMERA_BROADCAST,
    ("Sony", "Legacy 8-pin"): DeviceType.CAMERA_BROADCAST,
    # Z CAM
    ("Z CAM", "P2-R1"): DeviceType.CAMERA_PTZ,
}


def lookup_device_type(manufacturer: str, model: str) -> DeviceType | None:
    """Look up device type from the manual mapping."""
    for (mfr, substr), dtype in DEVICE_TYPE_MAP.items():
        if mfr != manufacturer:
            continue
        # Empty substring means "all models for this manufacturer"
        if substr == "" or substr in model:
            return dtype
    return None


# ─── Port parsing ───

def parse_ports(raw: str | None) -> list[str]:
    """Parse port column into canonical port names."""
    if not raw:
        return []
    parts = [p.strip() for p in str(raw).split(",")]
    result: list[str] = []
    for part in parts:
        port = _classify_port(part)
        if port and port not in result:
            result.append(port)
    return result


def _classify_port(text: str) -> str | None:
    """Classify a single port string."""
    t = text.strip()
    if not t:
        return None
    tl = t.lower()
    if "rs-422" in tl or "rs422" in tl:
        return "RS-422"
    if "rs-485" in tl or "rs485" in tl:
        return "RS-485"
    if "ethernet" in tl:
        return "Ethernet"
    if "usb" in tl:
        return "USB"
    if "lanc" in tl:
        return "LANC"
    if "sdi" in tl:
        return "SDI"
    if "hdmi" in tl:
        return "HDMI"
    if "wi-fi" in tl or "wifi" in tl:
        return "Wi-Fi"
    if "serial" in tl or "remote a" in tl or "8-pin" in tl or "visca serial" in tl:
        return "Serial"
    return None


# ─── Protocol parsing ───

# Order matters: check longer/more specific patterns first
PROTOCOL_RULES: list[tuple[str, str | list[str]]] = [
    ("VISCA IP", "VISCA over IP"),
    ("IP (Panasonic Native)", "Panasonic AW"),
    ("IP (REST API)", "REST API"),
    ("IP (through ATEM)", "Blackmagic IP"),
    ("Burano (Sony SDK)", "Sony SDK"),
    ("USB (Sony SDK)", "Sony SDK"),
    ("IP (Sony SDK)", "Sony SDK"),
    ("IP (s700 protocol)", "Sony RCP"),
    ("IP (FX9 or CBM)", "Sony RCP"),
    ("IP (Venice IP) or Serial (Venice Serial)", "Sony RCP"),
    ("IP or SDI", ["Blackmagic IP", "Blackmagic SDI"]),
    ("SDI (via Blackmagic Microconverter)", "Blackmagic SDI"),
    ("CBM (Sony Protocol)", "CBM"),
    ("CNS-Bridge", "CNS-Bridge"),
    ("USB (Panasonic Native)", "USB"),
    ("Serial (Panasonic Native)", "Serial"),
    ("Serial (Sony Protocol)", "Sony RCP"),
    ("Pelco RS-485", "Pelco-D"),
    ("ICPP", "ICPP"),
    ("Remote A", "Remote A"),
    ("XC", "Canon XC"),
]


def parse_protocols(raw: str | None) -> list[str]:
    """Parse protocol column into canonical protocol names."""
    if not raw:
        return []
    text = str(raw).strip()
    if not text:
        return []

    # Try exact/contains match against specific rules first
    for pattern, result in PROTOCOL_RULES:
        if pattern == text or pattern in text:
            if isinstance(result, list):
                return list(result)
            return [result]

    # Fallback: simple keyword matching
    tl = text.lower().strip()
    if tl == "visca":
        return ["VISCA"]
    if tl == "lanc":
        return ["LANC"]
    if tl == "sdi":
        return ["Blackmagic SDI"]
    if tl == "usb":
        return ["USB"]
    if tl == "serial":
        return ["Serial"]
    if tl == "ip":
        return ["HTTP API"]
    if "visca" in tl and "ip" in tl:
        return ["VISCA over IP"]

    # Return as-is if no rule matched
    return [text]


def build_notes(
    cable: str | None,
    notes_imp: str | None,
    notes_utiles: str | None,
) -> list[str]:
    """Combine cable, important notes, and useful notes into a list."""
    result: list[str] = []
    if cable:
        cable_str = str(cable).strip()
        if cable_str:
            result.append(f"Cable: {cable_str}")
    if notes_imp:
        imp_str = str(notes_imp).strip()
        if imp_str:
            result.append(imp_str)
    if notes_utiles:
        util_str = str(notes_utiles).strip()
        if util_str:
            result.append(util_str)
    return result


def main() -> None:
    """Import cameras from Excel into the database."""
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Loading workbook: {EXCEL_PATH}")
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        print("ERROR: No active worksheet found")
        sys.exit(1)

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        print("ERROR: Empty worksheet")
        sys.exit(1)

    header = rows[0]
    print(f"Header: {header}")
    print(f"Total data rows: {len(rows) - 1}")

    # Columns: Marque, Modèle, Surnom, Réf. Constructeur,
    #          Type de Port, Protocole, Câble Cyanview,
    #          Notes Importantes, Notes Utiles
    db = get_camera_database()

    imported = 0
    skipped_generic = 0
    skipped_empty = 0
    errors: list[str] = []
    mfr_counter: Counter[str] = Counter()
    dtype_counter: Counter[str] = Counter()

    for row_idx, row in enumerate(rows[1:], start=2):
        # Unpack columns (pad with None if row is short)
        padded = list(row) + [None] * (9 - len(row))
        marque = padded[0]
        modele = padded[1]
        surnom = padded[2]
        ref_constructeur = padded[3]
        type_port = padded[4]
        protocole = padded[5]
        cable = padded[6]
        notes_imp = padded[7]
        notes_utiles = padded[8]

        # Skip empty rows
        if not marque and not modele:
            skipped_empty += 1
            continue

        manufacturer = str(marque).strip() if marque else None
        model_name = str(modele).strip() if modele else None

        if not model_name:
            skipped_empty += 1
            continue

        # Skip Generic rows
        if manufacturer and manufacturer.lower() == "generic":
            skipped_generic += 1
            continue
        if model_name.lower() == "generic":
            skipped_generic += 1
            continue

        # Parse fields
        ports = parse_ports(str(type_port) if type_port else None)
        protocols = parse_protocols(str(protocole) if protocole else None)
        notes = build_notes(cable, notes_imp, notes_utiles)
        code_model = str(ref_constructeur).strip() if ref_constructeur else None
        nickname = str(surnom).strip() if surnom else None

        # Add nickname to notes if present
        if nickname:
            notes.insert(0, f"Also known as: {nickname}")

        # Determine device type via manual mapping
        device_type = lookup_device_type(manufacturer or "", model_name)

        camera = CameraCreate(
            name=model_name,
            manufacturer=manufacturer,
            code_model=code_model,
            device_type=device_type,
            ports=ports,
            protocols=protocols,
            notes=notes,
            confidence=1.0,
            needs_review=False,
        )

        # Normalize (will also infer device_type if still None)
        camera = normalize_camera_create(camera)

        try:
            result = db.upsert_camera(camera, CameraSource.MANUFACTURER)
            imported += 1
            mfr_name = manufacturer or "Unknown"
            mfr_counter[mfr_name] += 1
            dt_label = camera.device_type.value if camera.device_type else "unclassified"
            dtype_counter[dt_label] += 1
        except Exception as e:
            errors.append(f"Row {row_idx} ({model_name}): {e}")

    wb.close()

    # Rebuild FTS index
    print("\nRebuilding FTS index...")
    indexed = db.rebuild_fts_index()
    print(f"FTS index rebuilt: {indexed} cameras indexed")

    # Print stats
    print(f"\n{'='*60}")
    print("IMPORT RESULTS")
    print(f"{'='*60}")
    print(f"Imported:       {imported}")
    print(f"Skipped generic: {skipped_generic}")
    print(f"Skipped empty:   {skipped_empty}")
    print(f"Errors:          {len(errors)}")

    if errors:
        print("\nErrors:")
        for err in errors:
            print(f"  - {err}")

    print(f"\n--- By Manufacturer ({len(mfr_counter)} manufacturers) ---")
    for mfr, count in sorted(mfr_counter.items()):
        print(f"  {mfr:20s} {count:3d}")

    print(f"\n--- By Device Type ({len(dtype_counter)} types) ---")
    for dt, count in sorted(dtype_counter.items()):
        print(f"  {dt:25s} {count:3d}")

    # Integrity check
    print(f"\n{'='*60}")
    print("INTEGRITY CHECK")
    print(f"{'='*60}")

    # Clear cache to get fresh data after bulk import
    db._cache.invalidate()
    all_cameras = db.list_cameras()
    total = len(all_cameras)
    print(f"Total cameras in DB: {total}")

    no_type = sum(1 for c in all_cameras if not c.device_type)
    no_ports = sum(1 for c in all_cameras if not c.ports)
    no_protocols = sum(1 for c in all_cameras if not c.protocols)
    no_mfr = sum(1 for c in all_cameras if not c.manufacturer)

    print(f"Missing device_type: {no_type}")
    print(f"Missing ports:       {no_ports}")
    print(f"Missing protocols:   {no_protocols}")
    print(f"Missing manufacturer: {no_mfr}")

    if no_type > 0:
        print("\nCameras without device_type:")
        for c in all_cameras:
            if not c.device_type:
                print(f"  - {c.manufacturer} / {c.name}")

    # FTS test
    fts_results = db.search_cameras("Sony")
    print(f"\nFTS test 'Sony': {len(fts_results)} results")

    print(f"\n{'='*60}")
    print("IMPORT COMPLETE")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
